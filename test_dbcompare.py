from openpyxl import Workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
import allure
import pytest
import pyodbc

server = "10.0.0.14"
database="Company"
username="sa"
password="Password@123"


cnxn=pyodbc.connect('Driver={ODBC Driver 17 for SQL Server};\
                      SERVER='+server+';\
                      DATABASE='+database+';\
                      UID='+username+';\
                      PWD='+ password)

print(type(cnxn))
cursor = cnxn.cursor()
print(type(cursor))
sql_query = """
SELECT 
    e.emp_id,
    TRIM(UPPER(e.emp_name)) AS name,
    emp_dob AS dob,
    LOWER(emp_email) AS email,
    (CASE 
        WHEN Salary IS NULL THEN '0.0'
        WHEN Salary > '0' THEN Salary
     END) AS sal,
    d.dept_name,
    l.city,
    m.manager_name,
    CASE 
        WHEN Salary >= 0 AND Salary <= '29999' THEN 'F'
        WHEN Salary >= '30000' AND Salary <= '49999' THEN 'A'
        WHEN Salary >= '50000' AND Salary <= '69999' THEN 'B'
        WHEN Salary >= '70000' AND Salary <= '89999' THEN 'C'
        WHEN Salary >= '90000' AND Salary <= '109999' THEN 'D'
        WHEN Salary >= '110000' AND Salary <= '150000' THEN 'E'
    END AS Salary_grade,
    last_updated
FROM dbo.Employee e
LEFT JOIN dbo.Dept d 
    ON e.dept_id = d.dept_id AND e.dept_id IS NOT NULL
LEFT JOIN dbo.Location l 
    ON d.location_id = l.location_id AND l.location_id IS NOT NULL
LEFT JOIN dbo.Manager m 
    ON e.manager_id = m.manager_id AND m.manager_id IS NOT NULL
WHERE 
     Salary >= '0.00' OR Salary IS NULL
    AND emp_dob <= CAST(GETDATE() AS DATE)
    AND emp_email NOT LIKE '%_@__%.__%'
    AND emp_name IS NOT NULL
    AND Salary <= '150000'
"""
print("Connected to:", cnxn.getinfo(pyodbc.SQL_DATABASE_NAME))
sql_query2="""SELECT 
    e.emp_id,
    UPPER(LTRIM(RTRIM(e.emp_name))) AS name,
    e.emp_dob AS dob,
    LOWER(e.emp_email) AS email,
    e.salary,
    d.dept_name AS department,
    l.city,
    m.manager_name,
    CASE 
        WHEN e.salary BETWEEN 30000 AND 49999 THEN 'A'
        WHEN e.salary BETWEEN 50000 AND 69999 THEN 'B'
        WHEN e.salary BETWEEN 70000 AND 89999 THEN 'C'
        WHEN e.salary BETWEEN 90000 AND 109999 THEN 'D'
        WHEN e.salary BETWEEN 110000 AND 150000 THEN 'E'
        ELSE 'F'
    END AS salary_grade,
    e.last_updated
FROM 
    employee e
JOIN 
    Dept d ON e.dept_id = d.dept_id
JOIN 
    location l ON d.location_id = l.location_id
JOIN 
    manager m ON e.manager_id = m.manager_id
WHERE 
    e.salary IS NOT NULL
    AND e.salary >= 0
    AND e.emp_name IS NOT NULL
    AND e.emp_email IS NOT NULL
    AND e.emp_dob <= CAST(GETDATE() AS DATE)
    AND (
        e.salary BETWEEN 30000 AND 49999 OR
        e.salary BETWEEN 50000 AND 69999 OR
        e.salary BETWEEN 70000 AND 89999 OR
        e.salary BETWEEN 90000 AND 109999 OR
        e.salary BETWEEN 110000 AND 150000
    )"""
cursor.execute(sql_query)
#print(sql_query)

rows = cursor.fetchall()
# print(rows)
cursor.execute(sql_query2)
#print(sql_query)

rows_query2 = cursor.fetchall()
# print(rows_query2)
#call=sqlconection()
fill_style = PatternFill(start_color="FDD835", end_color="FDD835",fill_type="solid")
#TC1 of Table exists
@allure.feature("Checks Table existance ")
@allure.severity(allure.severity_level.BLOCKER)
def test_table_exists():
    table_query="SELECT * FROM information_schema.tables WHERE table_name ='Employee'"
    cursor.execute(table_query)
    result = cursor.fetchone()
    if result==1:
        # cursor.close()
        return True
    else:
        # cursor.close()
        return False

    allure.attach.file("table_exist.xlsx", name="Table exist", attachment_type=allure.attachment_type.XML)


@allure.feature("record count comparison")
@allure.severity(allure.severity_level.CRITICAL)
def test_record_Count_comparison():
    if(rows!=rows_query2):
        return True
    else:
        return False
    allure.attach.file("record_count.xlsx", name="Count match", attachment_type=allure.attachment_type.XML)

#TC2 of record count
@allure.feature("Database Comparison")
@allure.story("Compare two SQL queries and highlight differences")
@allure.severity(allure.severity_level.NORMAL)
def test_rec_compare_and_save():
    cursor.execute(sql_query)
    rows = cursor.fetchall()

    cursor.execute(sql_query2)
    rows_query2 = cursor.fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    fill_style = PatternFill(start_color="FDD835", end_color="FDD835", fill_type="solid")

    max_rows = max(len(rows), len(rows_query2))
    max_cols = max(len(rows[0]), len(rows_query2[0]))

    has_difference = False

    for j in range(max_cols):
        ws.cell(row=1, column=j + 1, value=f"Query1_Col{j+1}")
        ws.cell(row=1, column=j + 1 + max_cols, value=f"Query2_Col{j+1}")

    for i in range(max_rows):
        row1 = rows[i] if i < len(rows) else [None] * max_cols
        row2 = rows_query2[i] if i < len(rows_query2) else [None] * max_cols

        for j in range(max_cols):
            val1 = row1[j]
            val2 = row2[j]

            cell1 = ws.cell(row=i + 2, column=j + 1, value=val1)
            cell2 = ws.cell(row=i + 2, column=j + 1 + max_cols, value=val2)

            if val1 != val2:
                cell1.fill = fill_style
                cell2.fill = fill_style
                has_difference = True

    wb.save("data_Compare.xlsx")

    allure.attach.file("data_Compare.xlsx", name="Excel Comparison", attachment_type=allure.attachment_type.XML)

    assert not has_difference, "Differences found between query results"

